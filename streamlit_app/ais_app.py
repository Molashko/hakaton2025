import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import json
import uuid
import os
import sqlite3
import random
import time
import threading
import sys

# Добавляем путь к scripts для импорта Rule Engine
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'scripts'))
try:
    from rule_engine import RuleEngine
    RULE_ENGINE_AVAILABLE = True
except ImportError:
    RULE_ENGINE_AVAILABLE = False
    print("[WARN] Rule Engine не найден, используется простой алгоритм")

# Конфигурация страницы
st.set_page_config(
    page_title="СРЗ - Система распределения заявок",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS стили
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .section-header {
        font-size: 1.5rem;
        font-weight: bold;
        color: #2c3e50;
        margin-top: 2rem;
        margin-bottom: 1rem;
        border-bottom: 2px solid #3498db;
        padding-bottom: 0.5rem;
    }
    .metric-card {
        background-color: #f8f9fa;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #3498db;
        margin: 0.5rem 0;
    }
    .success-banner {
        background-color: #d4edda;
        color: #155724;
        padding: 1rem;
        border-radius: 0.5rem;
        border: 1px solid #c3e6cb;
        margin: 1rem 0;
        text-align: center;
        font-size: 1.1rem;
    }
</style>
""", unsafe_allow_html=True)

# Хранилище: SQLite
DB_PATH = os.environ.get('SQLITE_PATH', os.path.join(os.path.dirname(__file__), 'ais.db'))

def get_sqlite_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_sqlite():
    conn = get_sqlite_conn()
    cur = conn.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS tasks (
        id TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        category TEXT,
        priority TEXT,
        created_at TEXT,
        data TEXT
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS executors (
        id TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        email TEXT NOT NULL,
        department TEXT,
        skills TEXT,
        active INTEGER DEFAULT 1,
        daily_limit INTEGER DEFAULT 10,
        assigned_today INTEGER DEFAULT 0,
        created_at TEXT,
        data TEXT
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS assignments (
        id TEXT PRIMARY KEY,
        task_id TEXT,
        executor_id TEXT,
        assigned_at TEXT,
        score REAL
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS load_test_status (
        id INTEGER PRIMARY KEY,
        status TEXT,
        progress REAL,
        current INTEGER,
        total INTEGER,
        assigned INTEGER,
        elapsed REAL,
        performance REAL,
        message TEXT,
        updated_at TEXT
    )
    """)
    conn.commit()
    conn.close()

def get_load_test_status():
    """Получить статус нагрузочного теста из БД"""
    conn = get_sqlite_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM load_test_status WHERE id = 1")
    row = cur.fetchone()
    conn.close()
    if row:
        return {
            'status': row['status'],
            'progress': row['progress'] or 0,
            'current': row['current'] or 0,
            'total': row['total'] or 0,
            'assigned': row['assigned'] or 0,
            'elapsed': row['elapsed'] or 0,
            'performance': row['performance'] or 0,
            'message': row['message'] or ''
        }
    return None

def set_load_test_status(status, progress=0, current=0, total=0, assigned=0, elapsed=0, performance=0, message=''):
    """Установить статус нагрузочного теста в БД"""
    conn = get_sqlite_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO load_test_status (id, status, progress, current, total, assigned, elapsed, performance, message, updated_at)
        VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(id) DO UPDATE SET
            status=excluded.status,
            progress=excluded.progress,
            current=excluded.current,
            total=excluded.total,
            assigned=excluded.assigned,
            elapsed=excluded.elapsed,
            performance=excluded.performance,
            message=excluded.message,
            updated_at=excluded.updated_at
    """, (status, progress, current, total, assigned, elapsed, performance, message, datetime.now().isoformat()))
    conn.commit()
    conn.close()

def _json_dumps(obj):
    try:
        return json.dumps(obj, ensure_ascii=False)
    except Exception:
        return '{}'

def _json_loads(s):
    if not s:
        return {}
    try:
        return json.loads(s)
    except Exception:
        return {}

def load_tasks_from_db():
    conn = get_sqlite_conn()
    cur = conn.cursor()
    
    # Проверяем наличие колонки params
    cur.execute("PRAGMA table_info(tasks)")
    columns = [col[1] for col in cur.fetchall()]
    has_params = 'params' in columns
    
    if has_params:
        cur.execute("SELECT id,name,category,priority,created_at,data,params FROM tasks ORDER BY datetime(created_at) DESC")
    else:
        cur.execute("SELECT id,name,category,priority,created_at,data FROM tasks ORDER BY datetime(created_at) DESC")
    
    rows = cur.fetchall()
    conn.close()
    tasks = []
    for r in rows:
        t = {
            'id': r['id'],
            'name': r['name'],
            'category': r['category'],
            'priority': r['priority'],
            'created_at': r['created_at'] or datetime.now().isoformat()
        }
        
        # Загружаем дополнительные данные из data (для обратной совместимости)
        t.update(_json_loads(r['data']))
        
        # Загружаем params если колонка есть
        if has_params and r['params']:
            params = _json_loads(r['params'])
            if params:
                t['params'] = params
        
        tasks.append(t)
    return tasks

def load_executors_from_db():
    conn = get_sqlite_conn()
    cur = conn.cursor()
    
    # Проверяем наличие колонки params
    cur.execute("PRAGMA table_info(executors)")
    columns = [col[1] for col in cur.fetchall()]
    has_params = 'params' in columns
    
    if has_params:
        cur.execute("SELECT id,name,email,department,skills,active,daily_limit,assigned_today,created_at,data,params FROM executors ORDER BY name")
    else:
        cur.execute("SELECT id,name,email,department,skills,active,daily_limit,assigned_today,created_at,data FROM executors ORDER BY name")
    
    rows = cur.fetchall()
    conn.close()
    executors = []
    for r in rows:
        e = {
            'id': r['id'],
            'name': r['name'],
            'email': r['email'],
            'department': r['department'],
            'skills': r['skills'].split(',') if r['skills'] else [],
            'active': bool(r['active']),
            'daily_limit': r['daily_limit'],
            'assigned_today': r['assigned_today'],
            'created_at': r['created_at'] or datetime.now().isoformat()
        }
        
        # Загружаем дополнительные данные из data (для обратной совместимости)
        e.update(_json_loads(r['data']))
        
        # Загружаем params если колонка есть
        if has_params and r['params']:
            params = _json_loads(r['params'])
            if params:
                e['params'] = params
        
        executors.append(e)
    return executors

def load_assignments_from_db():
    conn = get_sqlite_conn()
    cur = conn.cursor()
    cur.execute("SELECT id,task_id,executor_id,assigned_at,score FROM assignments ORDER BY datetime(assigned_at) DESC")
    rows = cur.fetchall()
    conn.close()
    return [{
        'id': r['id'],
        'task_id': r['task_id'],
        'executor_id': r['executor_id'],
        'assigned_at': r['assigned_at'] or datetime.now().isoformat(),
        'score': r['score'] if r['score'] is not None else 0.0
    } for r in rows]

def save_task_to_db(task):
    conn = get_sqlite_conn()
    cur = conn.cursor()
    
    # Проверяем наличие колонки params
    cur.execute("PRAGMA table_info(tasks)")
    columns = [col[1] for col in cur.fetchall()]
    has_params = 'params' in columns
    
    base_keys = ['id','name','category','priority','created_at','params']
    data = {k: v for k, v in task.items() if k not in base_keys}
    
    # Получаем params
    params = task.get('params', {})
    params_json = _json_dumps(params) if params else '{}'
    
    if has_params:
        # Новая схема с колонкой params
    cur.execute("""
            INSERT INTO tasks(id,name,category,priority,created_at,data,params)
            VALUES(?,?,?,?,?,?,?)
            ON CONFLICT(id) DO UPDATE SET
                name=excluded.name,
                category=excluded.category,
                priority=excluded.priority,
                created_at=excluded.created_at,
                data=excluded.data,
                params=excluded.params
        """, (
            task['id'], task['name'], task.get('category',''), task.get('priority',''), 
            task['created_at'], _json_dumps(data), params_json
        ))
    else:
        # Старая схема без params
        cur.execute("""
            INSERT INTO tasks(id,name,category,priority,created_at,data)
        VALUES(?,?,?,?,?,?)
        ON CONFLICT(id) DO UPDATE SET
            name=excluded.name,
                category=excluded.category,
                priority=excluded.priority,
            created_at=excluded.created_at,
            data=excluded.data
    """, (
            task['id'], task['name'], task.get('category',''), task.get('priority',''), 
            task['created_at'], _json_dumps(data)
    ))
    
    conn.commit()
    conn.close()
    return True

def save_executor_to_db(executor):
    conn = get_sqlite_conn()
    cur = conn.cursor()
    
    # Проверяем наличие колонки params
    cur.execute("PRAGMA table_info(executors)")
    columns = [col[1] for col in cur.fetchall()]
    has_params = 'params' in columns
    
    base_keys = ['id','name','email','department','skills','active','daily_limit','assigned_today','created_at','params']
    data = {k: v for k, v in executor.items() if k not in base_keys}
    skills_str = ','.join(executor.get('skills', [])) if isinstance(executor.get('skills'), list) else executor.get('skills', '')
    
    # Получаем params
    params = executor.get('params', {})
    params_json = _json_dumps(params) if params else '{}'
    
    if has_params:
        # Новая схема с колонкой params
    cur.execute("""
            INSERT INTO executors(id,name,email,department,skills,active,daily_limit,assigned_today,created_at,data,params)
            VALUES(?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(id) DO UPDATE SET
            name=excluded.name,
            email=excluded.email,
                department=excluded.department,
                skills=excluded.skills,
            active=excluded.active,
                daily_limit=excluded.daily_limit,
                assigned_today=excluded.assigned_today,
                created_at=excluded.created_at,
                data=excluded.data,
                params=excluded.params
        """, (
            executor['id'], executor['name'], executor['email'], executor.get('department',''), skills_str, 
            1 if executor.get('active', True) else 0, executor.get('daily_limit', 10), 
            executor.get('assigned_today', 0), executor['created_at'], _json_dumps(data), params_json
        ))
    else:
        # Старая схема без params (для обратной совместимости)
        cur.execute("""
            INSERT INTO executors(id,name,email,department,skills,active,daily_limit,assigned_today,created_at,data)
            VALUES(?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(id) DO UPDATE SET
                name=excluded.name,
                email=excluded.email,
                department=excluded.department,
                skills=excluded.skills,
                active=excluded.active,
                daily_limit=excluded.daily_limit,
                assigned_today=excluded.assigned_today,
            created_at=excluded.created_at,
            data=excluded.data
    """, (
            executor['id'], executor['name'], executor['email'], executor.get('department',''), skills_str, 
            1 if executor.get('active', True) else 0, executor.get('daily_limit', 10), 
            executor.get('assigned_today', 0), executor['created_at'], _json_dumps(data)
    ))
    
    conn.commit()
    conn.close()
    return True

def save_assignment_to_db(assignment):
    conn = get_sqlite_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO assignments(id,task_id,executor_id,assigned_at,score)
        VALUES(?,?,?,?,?)
        ON CONFLICT(id) DO UPDATE SET
            task_id=excluded.task_id,
            executor_id=excluded.executor_id,
            assigned_at=excluded.assigned_at,
            score=excluded.score
    """, (
        assignment['id'], assignment['task_id'], assignment['executor_id'], assignment['assigned_at'], assignment.get('score', 0.0)
    ))
    conn.commit()
    conn.close()
    return True

def delete_executor_from_db(executor_id):
    conn = get_sqlite_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM assignments WHERE executor_id=?", (executor_id,))
    cur.execute("DELETE FROM executors WHERE id=?", (executor_id,))
    conn.commit()
    conn.close()
    return True

def clear_all_data_in_db():
    conn = get_sqlite_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM assignments")
    cur.execute("DELETE FROM tasks")
    cur.execute("UPDATE executors SET assigned_today=0")
    conn.commit()
    conn.close()
    return True

def reset_daily_counts_in_db():
    conn = get_sqlite_conn()
    cur = conn.cursor()
    cur.execute("UPDATE executors SET assigned_today=0")
    conn.commit()
    conn.close()
    return True

# Инициализация состояния сессии
def init_session_state():
    if 'db_initialized' not in st.session_state:
        init_sqlite()
        st.session_state.db_initialized = True
    if 'tasks' not in st.session_state:
        st.session_state.tasks = load_tasks_from_db()
    if 'executors' not in st.session_state:
        st.session_state.executors = load_executors_from_db()
    if 'assignments' not in st.session_state:
        st.session_state.assignments = load_assignments_from_db()

# Заголовок приложения
def render_header():
    st.markdown('<h1 class="main-header">⚖️ СРЗ - Система распределения заявок</h1>', unsafe_allow_html=True)
    st.markdown("""
    <div style="text-align: center; color: #7f8c8d; margin-bottom: 2rem;">
        Демонстрация интеллектуального распределения заявок между исполнителями<br>
        <strong>Справедливое распределение нагрузки в режиме реального времени</strong>
    </div>
    """, unsafe_allow_html=True)

# Главное меню
def render_main_menu():
    st.sidebar.markdown("## 📋 Главное меню")
    
    if 'current_page' not in st.session_state:
        st.session_state.current_page = "dashboard"
    
    current_page = st.session_state.current_page
    
    if st.sidebar.button("⚖️ Распределение", use_container_width=True, type="primary" if current_page == "dashboard" else "secondary"):
        st.session_state.current_page = "dashboard"
        st.rerun()
    
    if st.sidebar.button("👥 Исполнители", use_container_width=True, type="primary" if current_page == "executors" else "secondary"):
        st.session_state.current_page = "executors"
        st.rerun()
    
    if st.sidebar.button("🧪 Нагрузочное тестирование", use_container_width=True, type="primary" if current_page == "load_test" else "secondary"):
        st.session_state.current_page = "load_test"
        st.rerun()
    
    if st.sidebar.button("⚙️ Настройки", use_container_width=True, type="primary" if current_page == "settings" else "secondary"):
        st.session_state.current_page = "settings"
        st.rerun()
    
    return st.session_state.current_page

# Дашборд распределения
def _aggregate_per_minute(items, timestamp_key, window_minutes=5):
    now = datetime.now()
    start = now - timedelta(minutes=window_minutes)
    buckets = {}
    for i in range(window_minutes + 1):
        t = (start + timedelta(minutes=i)).replace(second=0, microsecond=0)
        buckets[t] = 0
    for it in items:
        try:
            ts = datetime.fromisoformat(it.get(timestamp_key, now.isoformat()))
            ts0 = ts.replace(second=0, microsecond=0)
            if ts0 >= start:
                buckets[ts0] = buckets.get(ts0, 0) + 1
        except Exception:
            pass
    times = sorted(buckets.keys())
    return pd.DataFrame({
        'Минута': [t.strftime('%H:%M') for t in times],
        'Количество': [buckets[t] for t in times],
    })

def export_dashboard_to_excel():
    """
    Экспортирует все метрики дашборда в Excel файл с диаграммами
    Возвращает байты файла для скачивания
    """
    from io import BytesIO
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter
    
    tasks = st.session_state.tasks
    executors = st.session_state.executors
    assignments = st.session_state.assignments
    active_executors = [e for e in executors if e.get('active', True)]
    
    # Создаем Excel writer
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # === Лист 1: Общие метрики ===
        if active_executors:
            total_assigned = sum(e['assigned_today'] for e in active_executors)
            avg_load = total_assigned / len(active_executors)
            
            # Расчет MAE
            utilizations = [e['assigned_today'] / e['daily_limit'] if e['daily_limit'] > 0 else 0 for e in active_executors]
            avg_util = sum(utilizations) / len(utilizations)
            mae = sum(abs(u - avg_util) for u in utilizations) / len(utilizations)
            
            # Статистика
            utilizations_pct = [u * 100 for u in utilizations]
            avg_utilization = sum(utilizations_pct) / len(utilizations_pct)
            std_utilization = pd.Series(utilizations_pct).std() if len(utilizations_pct) > 1 else 0
            min_utilization = min(utilizations_pct)
            max_utilization = max(utilizations_pct)
            range_utilization = max_utilization - min_utilization
        else:
            total_assigned = 0
            avg_load = 0
            mae = 0
            avg_utilization = 0
            std_utilization = 0
            min_utilization = 0
            max_utilization = 0
            range_utilization = 0
        
        recent_tasks = len([t for t in tasks if (datetime.now() - datetime.fromisoformat(t['created_at'])).seconds < 60])
        
        df_metrics = pd.DataFrame({
            'Метрика': [
                'Всего заявок в системе',
                'Новых заявок за последнюю минуту',
                'Активных исполнителей',
                'Обработано заявок',
                'Справедливость (MAE)',
                'Средняя нагрузка (заявок/исполнитель)',
                '---',
                'СТАТИСТИКА РАСПРЕДЕЛЕНИЯ:',
                'Средняя утилизация (%)',
                'Стандартное отклонение (σ, %)',
                'Минимальная утилизация (%)',
                'Максимальная утилизация (%)',
                'Разброс утилизации (%)',
            ],
            'Значение': [
                len(tasks),
                recent_tasks,
                len(active_executors),
                total_assigned,
                f'{mae:.3f}',
                f'{avg_load:.1f}',
                '',
                '',
                f'{avg_utilization:.1f}',
                f'{std_utilization:.2f}',
                f'{min_utilization:.1f}',
                f'{max_utilization:.1f}',
                f'{range_utilization:.1f}',
            ]
        })
        df_metrics.to_excel(writer, sheet_name='Общие метрики', index=False)
        
        # === Лист 2: Детали по исполнителям ===
        if active_executors:
            exec_data = []
            for e in active_executors:
                utilization = e['assigned_today'] / e['daily_limit'] if e['daily_limit'] > 0 else 0
                deviation = utilization * 100 - avg_utilization
                exec_data.append({
                    'ID': e['id'],
                    'Исполнитель': e['name'],
                    'Email': e['email'],
                    'Отдел': e.get('department', 'N/A'),
                    'Назначено заявок': e['assigned_today'],
                    'Дневной лимит': e['daily_limit'],
                    'Утилизация (%)': round(utilization * 100, 1),
                    'Отклонение от среднего (%)': round(deviation, 1),
                    'Навыки': ', '.join(e.get('skills', [])),
                    'Активен': 'Да' if e.get('active', True) else 'Нет'
                })
            df_executors = pd.DataFrame(exec_data)
            df_executors.to_excel(writer, sheet_name='Исполнители', index=False)
        
        # === Лист 3: Поступление заявок (последние 5 минут) ===
        df_tasks_min = _aggregate_per_minute(tasks, 'created_at', window_minutes=5)
        df_tasks_min.to_excel(writer, sheet_name='Поступление заявок', index=False)
        
        # === Лист 4: Назначения (последние 5 минут) ===
        df_assign_min = _aggregate_per_minute(assignments, 'assigned_at', window_minutes=5)
        df_assign_min.to_excel(writer, sheet_name='Назначения', index=False)
        
        # === Лист 5: Все назначения ===
        if assignments:
            assign_data = []
            for a in assignments:
                # Найдем исполнителя и заявку
                executor = next((e for e in executors if e['id'] == a['executor_id']), None)
                task = next((t for t in tasks if t['id'] == a['task_id']), None)
                
                assign_data.append({
                    'ID назначения': a['id'],
                    'ID заявки': a['task_id'],
                    'Категория заявки': task.get('category', 'N/A') if task else 'N/A',
                    'Приоритет': task.get('priority', 'N/A') if task else 'N/A',
                    'ID исполнителя': a['executor_id'],
                    'Имя исполнителя': executor['name'] if executor else 'N/A',
                    'Отдел': executor.get('department', 'N/A') if executor else 'N/A',
                    'Дата назначения': a['assigned_at'],
                    'Оценка (score)': round(a.get('score', 0), 2)
                })
            df_assignments = pd.DataFrame(assign_data)
            df_assignments.to_excel(writer, sheet_name='Все назначения', index=False)
        
        # === СОЗДАНИЕ ДИАГРАММ ===
        workbook = writer.book
        
        # === Диаграмма 1: Утилизация исполнителей (на листе "Исполнители") ===
        if active_executors and 'Исполнители' in workbook.sheetnames:
            ws_exec = workbook['Исполнители']
            
            # Столбчатая диаграмма утилизации
            chart1 = BarChart()
            chart1.title = "Утилизация исполнителей (%)"
            chart1.y_axis.title = "Утилизация (%)"
            chart1.x_axis.title = "Исполнитель"
            
            # Данные из колонки G (Утилизация)
            data = Reference(ws_exec, min_col=7, min_row=1, max_row=len(active_executors)+1)
            categories = Reference(ws_exec, min_col=2, min_row=2, max_row=len(active_executors)+1)
            
            chart1.add_data(data, titles_from_data=True)
            chart1.set_categories(categories)
            chart1.height = 12
            chart1.width = 20
            
            ws_exec.add_chart(chart1, "L2")
            
            # Диаграмма 2: Отклонение от среднего
            chart2 = BarChart()
            chart2.title = "Отклонение от среднего (%)"
            chart2.y_axis.title = "Отклонение (%)"
            chart2.x_axis.title = "Исполнитель"
            
            # Данные из колонки H (Отклонение от среднего)
            data2 = Reference(ws_exec, min_col=8, min_row=1, max_row=len(active_executors)+1)
            categories2 = Reference(ws_exec, min_col=2, min_row=2, max_row=len(active_executors)+1)
            
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(categories2)
            chart2.height = 12
            chart2.width = 20
            
            ws_exec.add_chart(chart2, "L22")
        
        # === Диаграмма 3: Поступление заявок (линейный график) ===
        if 'Поступление заявок' in workbook.sheetnames:
            ws_tasks = workbook['Поступление заявок']
            
            chart3 = LineChart()
            chart3.title = "Поступление заявок (последние 5 минут)"
            chart3.y_axis.title = "Количество заявок"
            chart3.x_axis.title = "Время"
            
            data3 = Reference(ws_tasks, min_col=2, min_row=1, max_row=ws_tasks.max_row)
            categories3 = Reference(ws_tasks, min_col=1, min_row=2, max_row=ws_tasks.max_row)
            
            chart3.add_data(data3, titles_from_data=True)
            chart3.set_categories(categories3)
            chart3.height = 12
            chart3.width = 20
            chart3.style = 10
            
            ws_tasks.add_chart(chart3, "E2")
        
        # === Диаграмма 4: Назначения (линейный график) ===
        if 'Назначения' in workbook.sheetnames:
            ws_assign = workbook['Назначения']
            
            chart4 = LineChart()
            chart4.title = "Назначения (последние 5 минут)"
            chart4.y_axis.title = "Количество назначений"
            chart4.x_axis.title = "Время"
            
            data4 = Reference(ws_assign, min_col=2, min_row=1, max_row=ws_assign.max_row)
            categories4 = Reference(ws_assign, min_col=1, min_row=2, max_row=ws_assign.max_row)
            
            chart4.add_data(data4, titles_from_data=True)
            chart4.set_categories(categories4)
            chart4.height = 12
            chart4.width = 20
            chart4.style = 12
            
            ws_assign.add_chart(chart4, "E2")
        
        # === Диаграмма 5: Распределение по отделам (круговая) ===
        if active_executors and 'Исполнители' in workbook.sheetnames:
            ws_exec = workbook['Исполнители']
            
            # Создаем сводку по отделам
            departments = {}
            for e in active_executors:
                dept = e.get('department', 'N/A')
                departments[dept] = departments.get(dept, 0) + e['assigned_today']
            
            # Добавляем данные для круговой диаграммы справа от таблицы
            start_row = len(active_executors) + 5
            ws_exec.cell(start_row, 11, "Отдел")
            ws_exec.cell(start_row, 12, "Заявок")
            
            row = start_row + 1
            for dept, count in departments.items():
                ws_exec.cell(row, 11, dept)
                ws_exec.cell(row, 12, count)
                row += 1
            
            # Создаем круговую диаграмму
            chart5 = PieChart()
            chart5.title = "Распределение заявок по отделам"
            
            labels = Reference(ws_exec, min_col=11, min_row=start_row+1, max_row=start_row+len(departments))
            data5 = Reference(ws_exec, min_col=12, min_row=start_row, max_row=start_row+len(departments))
            
            chart5.add_data(data5, titles_from_data=True)
            chart5.set_categories(labels)
            chart5.height = 12
            chart5.width = 15
            
            # Добавляем подписи данных
            chart5.dataLabels = DataLabelList()
            chart5.dataLabels.showPercent = True
            
            ws_exec.add_chart(chart5, "L42")
    
    output.seek(0)
    return output.getvalue()

def render_dashboard():
    st.markdown('<h2 class="section-header">⚖️ Распределение заявок</h2>', unsafe_allow_html=True)
    
    # Кнопка экспорта в Excel
    col_export1, col_export2, col_export3 = st.columns([1, 1, 4])
    with col_export1:
        if st.button("📥 Экспорт в Excel", use_container_width=True, type="primary"):
            try:
                excel_data = export_dashboard_to_excel()
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"распределение_заявок_{timestamp}.xlsx"
                
                st.download_button(
                    label="⬇️ Скачать файл",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.success(f"✅ Excel файл готов к скачиванию!")
            except Exception as e:
                st.error(f"❌ Ошибка при создании Excel: {str(e)}")
    
    with col_export2:
        if st.button("🔄 Обновить данные", use_container_width=True):
                    st.session_state.tasks = load_tasks_from_db()
            st.session_state.executors = load_executors_from_db()
            st.session_state.assignments = load_assignments_from_db()
                    st.rerun()
    
    st.markdown("---")

    # Показываем индикатор нагрузочного тестирования если оно запущено
    test_status_data = get_load_test_status()
    if test_status_data and test_status_data['status'] == 'running':
        progress = test_status_data['progress']
        current = test_status_data['current']
        total = test_status_data['total']
        assigned = test_status_data['assigned']
        
        with st.expander("🔄 Нагрузочное тестирование в процессе...", expanded=True):
            st.progress(progress)
            st.write(f"**Обработано:** {current}/{total} заявок | **Назначено:** {assigned}")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("⏹️ Остановить"):
                    set_load_test_status('stopped')
                    st.rerun()
        with col2:
                if st.button("🧪 Перейти к тестированию"):
                    st.session_state.current_page = "load_test"
                st.rerun()
        
    # Автообновление
    auto_refresh_enabled = st.session_state.get('auto_refresh', True)
    if auto_refresh_enabled:
                st.session_state.tasks = load_tasks_from_db()
        st.session_state.executors = load_executors_from_db()
                st.session_state.assignments = load_assignments_from_db()
        
        # Автоматическое распределение нераспределенных заявок
        assigned_count = auto_assign_unassigned_tasks()
        if assigned_count > 0:
            # Обновляем данные после распределения
            st.session_state.tasks = load_tasks_from_db()
            st.session_state.executors = load_executors_from_db()
            st.session_state.assignments = load_assignments_from_db()
        
        try:
            from streamlit_autorefresh import st_autorefresh
            st_autorefresh(interval=2000, limit=None, key="ais_realtime")
        except Exception:
            st.markdown('<meta http-equiv="refresh" content="2">', unsafe_allow_html=True)
    
    # Индикатор автоматического распределения
    tasks = st.session_state.tasks
    assignments = st.session_state.assignments
    assigned_task_ids = set(a['task_id'] for a in assignments)
    unassigned_count = len([t for t in tasks if t['id'] not in assigned_task_ids])
    
    if unassigned_count > 0:
        st.info(f"⚠️ Нераспределенных заявок: **{unassigned_count}** (автоматически распределяются)")
        st.markdown("---")
    elif len(assignments) > 0:
        st.success(f"✅ Все заявки распределены! Всего назначений: {len(assignments)}")
        st.markdown("---")
    
    # Ключевые метрики
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        recent_tasks = len([t for t in st.session_state.tasks if (datetime.now() - datetime.fromisoformat(t['created_at'])).seconds < 60])
        st.metric(
            label="📊 Всего заявок в системе",
            value=len(st.session_state.tasks),
            delta=f"+{recent_tasks} за последнюю минуту" if recent_tasks > 0 else "Нет новых"
        )
    
    with col2:
        active_executors = [e for e in st.session_state.executors if e.get('active', True)]
        total_assigned = sum(e['assigned_today'] for e in active_executors)
        st.metric(
            label="👥 Активных исполнителей",
            value=len(active_executors),
            delta=f"Обработано заявок: {total_assigned}",
            delta_color="off"
        )
    
    with col3:
        # Расчет MAE (Mean Absolute Error) для справедливости
        mae = 0
        if active_executors:
            utilizations = [e['assigned_today'] / e['daily_limit'] if e['daily_limit'] > 0 else 0 for e in active_executors]
            avg_util = sum(utilizations) / len(utilizations)
            mae = sum(abs(u - avg_util) for u in utilizations) / len(utilizations)
            st.metric(
                label="⚖️ Справедливость (MAE)",
                value=f"{mae:.3f}",
                delta="Чем ближе к 0, тем лучше",
                delta_color="inverse"
            )
        else:
            st.metric(label="⚖️ Справедливость (MAE)", value="N/A")
    
    with col4:
        if active_executors:
            avg_load = sum(e['assigned_today'] for e in active_executors) / len(active_executors)
            st.metric(
                label="📈 Средняя нагрузка",
                value=f"{avg_load:.1f}",
                delta=f"заявок/исполнитель"
            )
                        else:
            st.metric(label="📈 Средняя нагрузка", value="0")
    
    # Графики распределения
    st.markdown("### 📊 Распределение нагрузки")
    
    if active_executors:
        # График утилизации исполнителей
        exec_data = []
        for e in active_executors:
            utilization = e['assigned_today'] / e['daily_limit'] if e['daily_limit'] > 0 else 0
            exec_data.append({
                'Исполнитель': e['name'],
                'Назначено': e['assigned_today'],
                'Лимит': e['daily_limit'],
                'Утилизация': utilization * 100
            })
        
        df_exec = pd.DataFrame(exec_data)
        
        # Расчет статистики
        avg_utilization = df_exec['Утилизация'].mean()
        std_utilization = df_exec['Утилизация'].std() if len(df_exec) > 1 else 0
        min_utilization = df_exec['Утилизация'].min()
        max_utilization = df_exec['Утилизация'].max()
        range_utilization = max_utilization - min_utilization
        
        # Статистические метрики распределения
        st.markdown("#### 📈 Статистика распределения")
        stat_col1, stat_col2, stat_col3 = st.columns(3)
        
        with stat_col1:
            st.metric(
                label="📊 Среднее",
                value=f"{avg_utilization:.1f}%",
                help="Средняя утилизация всех исполнителей"
            )
        
        with stat_col2:
            st.metric(
                label="📏 Станд. отклонение",
                value=f"{std_utilization:.2f}%",
                delta="Чем меньше, тем лучше" if std_utilization > 0 else "Идеально!",
                delta_color="inverse",
                help="Разброс утилизации (σ). Хорошо: < 5%"
            )
        
        with stat_col3:
            st.metric(
                label="📐 Разброс",
                value=f"{range_utilization:.1f}%",
                delta=f"Min: {min_utilization:.1f}%, Max: {max_utilization:.1f}%",
                help="Разница между максимальной и минимальной утилизацией"
            )
        
        st.markdown("---")
        
        # График 1: Утилизация с линией среднего
        fig1 = px.bar(
            df_exec,
            x='Исполнитель',
            y='Утилизация',
            color='Утилизация',
            color_continuous_scale=['green', 'yellow', 'orange', 'red'],
            range_color=[0, 100],
            title="Утилизация исполнителей (%)",
            labels={'Утилизация': 'Утилизация (%)'}
        )
        fig1.add_hline(y=100, line_dash="dash", line_color="red", annotation_text="Лимит (100%)", annotation_position="right")
        fig1.add_hline(y=avg_utilization, line_dash="dot", line_color="blue", 
                      annotation_text=f"Среднее: {avg_utilization:.1f}%", 
                      annotation_position="left")
        fig1.update_layout(xaxis_tickangle=-45, showlegend=False, height=400)
        st.plotly_chart(fig1, use_container_width=True)
        
        # График 2 и 3 в двух колонках
        col_graph1, col_graph2 = st.columns(2)
        
        with col_graph1:
            # График отклонений от среднего
            df_exec['Отклонение'] = df_exec['Утилизация'] - avg_utilization
            
            fig2 = px.bar(
                df_exec,
                x='Исполнитель',
                y='Отклонение',
                color='Отклонение',
                color_continuous_scale=['red', 'yellow', 'green', 'yellow', 'red'],
                color_continuous_midpoint=0,
                title="Отклонение от среднего (%)",
                labels={'Отклонение': 'Отклонение (%)'}
            )
            fig2.add_hline(y=0, line_dash="solid", line_color="gray", line_width=1)
            fig2.update_layout(xaxis_tickangle=-45, showlegend=False, height=350)
            st.plotly_chart(fig2, use_container_width=True)
        
        with col_graph2:
            # Гистограмма распределения утилизации
            fig3 = px.histogram(
                df_exec,
                x='Утилизация',
                nbins=min(10, len(df_exec)),
                title="Гистограмма распределения утилизации",
                labels={'Утилизация': 'Утилизация (%)', 'count': 'Количество исполнителей'},
                color_discrete_sequence=['#1f77b4']
            )
            fig3.add_vline(x=avg_utilization, line_dash="dash", line_color="red",
                          annotation_text=f"Среднее: {avg_utilization:.1f}%",
                          annotation_position="top")
            fig3.update_layout(showlegend=False, height=350, bargap=0.1)
            st.plotly_chart(fig3, use_container_width=True)
        
        # Детальная таблица
        st.markdown("### 📋 Детали распределения")
        df_display = df_exec.copy()
        df_display['Утилизация'] = df_display['Утилизация'].apply(lambda x: f"{x:.1f}%")
        df_display['Отклонение'] = df_display['Отклонение'].apply(lambda x: f"{x:+.1f}%")
        df_display = df_display[['Исполнитель', 'Назначено', 'Лимит', 'Утилизация', 'Отклонение']]
        st.dataframe(df_display, use_container_width=True, hide_index=True)
    else:
        st.info("👥 Нет активных исполнителей. Добавьте исполнителей в разделе 'Исполнители'")
    
    # Графики активности
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### ⏱ Поступление заявок (последние 5 минут)")
        df_tasks_min = _aggregate_per_minute(st.session_state.tasks, 'created_at', window_minutes=5)
        fig2 = px.line(df_tasks_min, x='Минута', y='Количество', markers=True, color_discrete_sequence=['#1f77b4'])
        st.plotly_chart(fig2, use_container_width=True)
    
    with col2:
        st.markdown("### ⚡ Назначения (последние 5 минут)")
        df_assign_min = _aggregate_per_minute(st.session_state.assignments, 'assigned_at', window_minutes=5)
        fig3 = px.line(df_assign_min, x='Минута', y='Количество', markers=True, color_discrete_sequence=['#2ca02c'])
        st.plotly_chart(fig3, use_container_width=True)

# Управление исполнителями
def render_executors_management():
    st.markdown('<h2 class="section-header">👥 Управление исполнителями</h2>', unsafe_allow_html=True)
    
    # Проверка на редактирование исполнителя
    editing_executor_id = None
    for executor in st.session_state.executors:
        if st.session_state.get(f"editing_executor_{executor['id']}", False):
            editing_executor_id = executor['id']
            break
    
    if editing_executor_id:
        # Редактирование существующего исполнителя
        executor_to_edit = next(e for e in st.session_state.executors if e['id'] == editing_executor_id)
        st.markdown("### ✏️ Редактировать исполнителя")
        
        col1, col2 = st.columns(2)
        
        with col1:
            executor_name = st.text_input("Имя исполнителя", value=executor_to_edit['name'], key="edit_executor_name")
            executor_email = st.text_input("Email", value=executor_to_edit['email'], key="edit_executor_email")
            department = st.selectbox("Отдел", ["IT", "Строительство", "Страхование", "Консалтинг", "Другое"], 
                                     index=["IT", "Строительство", "Страхование", "Консалтинг", "Другое"].index(executor_to_edit.get('department', 'IT')), 
                                     key="edit_department")
        
        with col2:
            skills = st.multiselect("Навыки", ["Python", "JavaScript", "Строительство", "Проектирование", "Анализ рисков", "Продажи"], 
                                   default=executor_to_edit.get('skills', []), key="edit_skills")
            daily_limit = st.number_input("Дневной лимит заявок", min_value=1, max_value=100, value=executor_to_edit.get('daily_limit', 10), key="edit_daily_limit")
            active = st.checkbox("Активен", value=executor_to_edit.get('active', True), key="edit_active")
        
        # Дополнительные параметры
        st.markdown("### 🎯 Дополнительные параметры")
        
        with st.expander("📝 Управление параметрами", expanded=True):
            # Получаем текущие параметры
            current_params = executor_to_edit.get('params', {})
            if isinstance(current_params, str):
                try:
                    current_params = json.loads(current_params)
                except:
                    current_params = {}
            
            # Инициализируем в session_state если нет
            if f"edit_params_{editing_executor_id}" not in st.session_state:
                st.session_state[f"edit_params_{editing_executor_id}"] = current_params.copy()
            
            params = st.session_state[f"edit_params_{editing_executor_id}"]
            
            # Добавление нового параметра
            st.markdown("**Добавить новый параметр:**")
            col_p1, col_p2, col_p3, col_p4 = st.columns([2, 2, 2, 1])
            
            with col_p1:
                new_param_key = st.text_input("Ключ", key=f"edit_new_param_key_{editing_executor_id}", 
                                             placeholder="например: experience_years")
            
            with col_p2:
                param_type = st.selectbox("Тип", ["Текст", "Число", "Список"], key=f"edit_param_type_{editing_executor_id}")
            
            with col_p3:
                if param_type == "Текст":
                    new_param_value = st.text_input("Значение", key=f"edit_new_param_value_{editing_executor_id}",
                                                   placeholder="например: Senior")
                elif param_type == "Число":
                    new_param_value = st.number_input("Значение", key=f"edit_new_param_value_num_{editing_executor_id}",
                                                     value=0)
                else:  # Список
                    new_param_value_str = st.text_input("Значения (через запятую)", 
                                                        key=f"edit_new_param_value_list_{editing_executor_id}",
                                                        placeholder="например: AWS, Python")
                    new_param_value = [v.strip() for v in new_param_value_str.split(',') if v.strip()] if new_param_value_str else []
            
            with col_p4:
                st.write("")
                st.write("")
                if st.button("➕", key=f"edit_add_param_{editing_executor_id}", help="Добавить параметр"):
                    if new_param_key and new_param_key not in params:
                        params[new_param_key] = new_param_value
                        st.session_state[f"edit_params_{editing_executor_id}"] = params
                        st.rerun()
                    elif new_param_key in params:
                        st.error(f"Параметр '{new_param_key}' уже существует!")
            
            # Отображение существующих параметров
            if params:
                st.markdown("---")
                st.markdown("**Текущие параметры:**")
                
                params_to_delete = []
                for key, value in params.items():
                    col_k, col_v, col_d = st.columns([2, 4, 1])
                    
                    with col_k:
                        st.markdown(f"**{key}:**")
                    
                    with col_v:
                        # Определяем тип значения
                        if isinstance(value, list):
                            new_value = st.text_input(
                                f"value_{key}", 
                                value=", ".join(str(v) for v in value),
                                key=f"edit_param_edit_{editing_executor_id}_{key}",
                                label_visibility="collapsed"
                            )
                            params[key] = [v.strip() for v in new_value.split(',') if v.strip()]
                        elif isinstance(value, (int, float)):
                            params[key] = st.number_input(
                                f"value_{key}",
                                value=float(value),
                                key=f"edit_param_edit_{editing_executor_id}_{key}",
                                label_visibility="collapsed"
                            )
                        else:
                            params[key] = st.text_input(
                                f"value_{key}",
                                value=str(value),
                                key=f"edit_param_edit_{editing_executor_id}_{key}",
                                label_visibility="collapsed"
                            )
                    
                    with col_d:
                        if st.button("🗑️", key=f"edit_delete_param_{editing_executor_id}_{key}", help=f"Удалить {key}"):
                            params_to_delete.append(key)
                
                # Удаляем помеченные параметры
                for key in params_to_delete:
                    del params[key]
                    st.session_state[f"edit_params_{editing_executor_id}"] = params
                    st.rerun()
            else:
                st.info("📝 Параметры не добавлены. Добавьте первый параметр выше.")
        
        st.markdown("---")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("💾 Сохранить изменения", type="primary"):
                if executor_name and executor_email:
                    executor_to_edit['name'] = executor_name
                    executor_to_edit['email'] = executor_email
                    executor_to_edit['department'] = department
                    executor_to_edit['skills'] = skills
                    executor_to_edit['daily_limit'] = daily_limit
                    executor_to_edit['active'] = active
                    
                    # Сохраняем параметры
                    if f"edit_params_{editing_executor_id}" in st.session_state:
                        executor_to_edit['params'] = st.session_state[f"edit_params_{editing_executor_id}"]
                    
                    save_executor_to_db(executor_to_edit)
                    
                    # Очищаем session_state
                    if f"edit_params_{editing_executor_id}" in st.session_state:
                        del st.session_state[f"edit_params_{editing_executor_id}"]
                    st.session_state[f"editing_executor_{editing_executor_id}"] = False
                    st.session_state.executors = load_executors_from_db()
                    
                    # Автоматически распределяем нераспределенные заявки
                    assigned_count = auto_assign_unassigned_tasks()
                    
                    # Обновляем данные в сессии
                    st.session_state.assignments = load_assignments_from_db()
                    st.session_state.executors = load_executors_from_db()
                    
                    if assigned_count > 0:
                        st.success(f"✅ Исполнитель обновлен! Автоматически назначено заявок: {assigned_count}")
                    else:
                    st.success("✅ Исполнитель успешно обновлен!")
                    st.rerun()
                else:
                    st.error("❌ Заполните обязательные поля")
        
        with col2:
            if st.button("❌ Отменить", type="secondary"):
                st.session_state[f"editing_executor_{editing_executor_id}"] = False
                # Очищаем session_state параметров
                if f"edit_params_{editing_executor_id}" in st.session_state:
                    del st.session_state[f"edit_params_{editing_executor_id}"]
                st.rerun()
        
        with col3:
            if st.button("🗑️ Удалить исполнителя", type="secondary"):
                delete_executor_from_db(editing_executor_id)
                st.session_state.executors = load_executors_from_db()
                st.session_state.assignments = load_assignments_from_db()
                st.session_state[f"editing_executor_{editing_executor_id}"] = False
                st.success("✅ Исполнитель удален!")
                st.rerun()
        
        return
    
    # Создание нового исполнителя
    st.markdown("### ➕ Добавить нового исполнителя")
    
    col1, col2 = st.columns(2)
    
    with col1:
        executor_name = st.text_input("Имя исполнителя", placeholder="Введите имя исполнителя")
        executor_email = st.text_input("Email", placeholder="email@example.com")
        department = st.selectbox("Отдел", ["IT", "Строительство", "Страхование", "Консалтинг", "Другое"])
    
    with col2:
        skills = st.multiselect("Навыки", ["Python", "JavaScript", "Строительство", "Проектирование", "Анализ рисков", "Продажи"])
        daily_limit = st.number_input("Дневной лимит заявок", min_value=1, max_value=100, value=10)
        active = st.checkbox("Активен", value=True)
    
    # Дополнительные параметры для нового исполнителя
    st.markdown("### 🎯 Дополнительные параметры")
    
    with st.expander("📝 Управление параметрами", expanded=False):
        # Инициализируем в session_state если нет
        if "new_executor_params" not in st.session_state:
            st.session_state.new_executor_params = {}
        
        params = st.session_state.new_executor_params
        
        # Добавление нового параметра
        st.markdown("**Добавить новый параметр:**")
        col_p1, col_p2, col_p3, col_p4 = st.columns([2, 2, 2, 1])
        
        with col_p1:
            new_param_key = st.text_input("Ключ", key="new_exec_param_key", 
                                         placeholder="например: experience_years")
        
        with col_p2:
            param_type = st.selectbox("Тип", ["Текст", "Число", "Список"], key="new_exec_param_type")
        
        with col_p3:
            if param_type == "Текст":
                new_param_value = st.text_input("Значение", key="new_exec_param_value",
                                               placeholder="например: Senior")
            elif param_type == "Число":
                new_param_value = st.number_input("Значение", key="new_exec_param_value_num",
                                                 value=0)
            else:  # Список
                new_param_value_str = st.text_input("Значения (через запятую)", 
                                                    key="new_exec_param_value_list",
                                                    placeholder="например: AWS, Python")
                new_param_value = [v.strip() for v in new_param_value_str.split(',') if v.strip()] if new_param_value_str else []
        
        with col_p4:
            st.write("")
            st.write("")
            if st.button("➕", key="new_exec_add_param", help="Добавить параметр"):
                if new_param_key and new_param_key not in params:
                    params[new_param_key] = new_param_value
                    st.session_state.new_executor_params = params
                    st.rerun()
                elif new_param_key in params:
                    st.error(f"Параметр '{new_param_key}' уже существует!")
        
        # Отображение существующих параметров
        if params:
            st.markdown("---")
            st.markdown("**Текущие параметры:**")
            
            params_to_delete = []
            for key, value in params.items():
                col_k, col_v, col_d = st.columns([2, 4, 1])
                
                with col_k:
                    st.markdown(f"**{key}:**")
                
                with col_v:
                    # Определяем тип значения
                    if isinstance(value, list):
                        new_value = st.text_input(
                            f"value_{key}", 
                            value=", ".join(str(v) for v in value),
                            key=f"new_exec_param_edit_{key}",
                            label_visibility="collapsed"
                        )
                        params[key] = [v.strip() for v in new_value.split(',') if v.strip()]
                    elif isinstance(value, (int, float)):
                        params[key] = st.number_input(
                            f"value_{key}",
                            value=float(value),
                            key=f"new_exec_param_edit_{key}",
                            label_visibility="collapsed"
                        )
                        else:
                        params[key] = st.text_input(
                            f"value_{key}",
                            value=str(value),
                            key=f"new_exec_param_edit_{key}",
                            label_visibility="collapsed"
                        )
                
                with col_d:
                    if st.button("🗑️", key=f"new_exec_delete_param_{key}", help=f"Удалить {key}"):
                        params_to_delete.append(key)
            
            # Удаляем помеченные параметры
            for key in params_to_delete:
                del params[key]
                st.session_state.new_executor_params = params
                st.rerun()
                    else:
            st.info("📝 Параметры не добавлены. Добавьте первый параметр выше.")
    
    st.markdown("---")
    
    if st.button("👥 Добавить исполнителя", type="primary"):
        if executor_name and executor_email:
            new_executor = {
                'id': str(uuid.uuid4()),
                'name': executor_name,
                'email': executor_email,
                'department': department,
                'skills': skills,
                'active': active,
                'daily_limit': daily_limit,
                'assigned_today': 0,
                'created_at': datetime.now().isoformat()
            }
            
            # Добавляем параметры если есть
            if st.session_state.new_executor_params:
                new_executor['params'] = st.session_state.new_executor_params.copy()
            
            save_executor_to_db(new_executor)
            
            # Очищаем session_state параметров
            st.session_state.new_executor_params = {}
            
            st.session_state.executors = load_executors_from_db()
            
            # Автоматически распределяем нераспределенные заявки
            assigned_count = auto_assign_unassigned_tasks()
            
            # Обновляем данные в сессии
            st.session_state.assignments = load_assignments_from_db()
            st.session_state.executors = load_executors_from_db()
            
            if assigned_count > 0:
                st.success(f"✅ Исполнитель успешно добавлен! Автоматически назначено заявок: {assigned_count}")
            else:
            st.success("✅ Исполнитель успешно добавлен!")
            st.rerun()
        else:
            st.error("❌ Заполните обязательные поля")
    
    # Список исполнителей
    st.markdown("### 👥 Список исполнителей")
    
    if st.session_state.executors:
        for executor in st.session_state.executors:
            with st.container():
                col1, col2, col3 = st.columns([3, 1, 1])
                
                with col1:
                    status_color = "🟢" if executor.get('active', True) else "🔴"
                    st.markdown(f"**{status_color} {executor['name']}**")
                    st.markdown(f"**Email:** {executor['email']} | **Отдел:** {executor['department']}")
                    st.markdown(f"**Навыки:** {', '.join(executor['skills'])} | **Назначено сегодня:** {executor['assigned_today']}/{executor['daily_limit']}")
                    
                    # Отображаем дополнительные параметры если есть
                    params = executor.get('params', {})
                    if isinstance(params, str):
                        try:
                            params = json.loads(params)
                        except:
                            params = {}
                    
                    if params:
                        params_str = " | ".join([f"{k}: {v}" for k, v in list(params.items())[:3]])
                        if len(params) > 3:
                            params_str += f" | +{len(params)-3} еще"
                        st.markdown(f"**📝 Параметры:** {params_str}")
                
                with col2:
                    if st.button(f"✏️ Редактировать", key=f"edit_exec_{executor['id']}"):
                        st.session_state[f"editing_executor_{executor['id']}"] = True
                        st.rerun()
                
                with col3:
                    if st.button(f"🗑️ Удалить", key=f"delete_exec_{executor['id']}"):
                        delete_executor_from_db(executor['id'])
                        st.session_state.executors = load_executors_from_db()
                        st.session_state.assignments = load_assignments_from_db()
                        st.success("✅ Исполнитель удален!")
                        st.rerun()
                
                st.markdown("---")
    else:
        st.info("👥 Исполнителей пока нет. Добавьте первого исполнителя выше.")

# Нагрузочное тестирование
def load_rule_engine():
    """Загрузить Rule Engine из конфигурации"""
    config_path = os.path.join(os.path.dirname(__file__), '..', 'config', 'matching_rules.json')
    
    if not os.path.exists(config_path):
        print(f"[WARN] Конфигурация правил не найдена: {config_path}")
        return None
    
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        engine = RuleEngine(config)
        print(f"[OK] Rule Engine загружен, правил: {len(config.get('rules', []))}")
        return engine
    except Exception as e:
        print(f"[ERROR] Ошибка загрузки Rule Engine: {e}")
        return None


def find_best_executor_simple(task, executors):
    """
    Алгоритм поиска лучшего исполнителя
    Поддерживает:
    - Простой алгоритм (если Rule Engine недоступен)
    - Rule Engine (если доступен и настроен)
    """
    # Фильтруем только активных исполнителей с доступными слотами
    active_executors = [e for e in executors if e.get('active', True) and e['assigned_today'] < e['daily_limit']]
    if not active_executors:
        return None
    
    # Пробуем использовать Rule Engine
    if RULE_ENGINE_AVAILABLE:
        try:
            engine = load_rule_engine()
            if engine:
                # Подготовка данных для Rule Engine
                # Обогащаем данные для использования в формулах
                for executor in active_executors:
                    executor['assigned_count'] = executor.get('assigned_today', 0)
                    executor['max_assignments'] = executor.get('daily_limit', 10)
                    # Добавляем params из data если есть
                    if 'params' not in executor and 'data' in executor:
                        executor['params'] = executor.get('data', {})
                
                # Добавляем params к заявке
                if 'params' not in task and 'data' in task:
                    task['params'] = task.get('data', {})
                
                # Добавляем is_active для правила active_executor
                task['is_active'] = 1
                
                # Используем Rule Engine
                result = engine.find_best_match(task, active_executors)
                
                if result:
                    executor, score, matched_rules = result
                    print(f"[Rule Engine] Score: {score:.2f}, Правил: {len(matched_rules)}")
                    return executor, score
        except Exception as e:
            print(f"[WARN] Rule Engine error: {e}, fallback to simple algorithm")
    
    # Fallback: простой алгоритм
    best_executor = None
    best_score = -1.0
    
    for executor in active_executors:
        # Расчет score на основе утилизации
        utilization = executor['assigned_today'] / executor['daily_limit'] if executor['daily_limit'] > 0 else 0
        fairness_score = 1.0 - utilization
        
        # Бонус за совпадение отдела
        department_bonus = 1.5 if executor.get('department') == task.get('category') else 1.0
        
        # Бонус за приоритет
        priority_bonus = {'Критический': 2.0, 'Высокий': 1.5, 'Средний': 1.0, 'Низкий': 0.8}.get(task.get('priority', 'Средний'), 1.0)
        
        final_score = fairness_score * department_bonus * priority_bonus
        
        if final_score > best_score:
            best_score = final_score
            best_executor = executor
    
    return best_executor, best_score if best_executor else None

def auto_assign_unassigned_tasks():
    """Автоматически распределяет все нераспределенные заявки"""
    tasks = load_tasks_from_db()
    assignments = load_assignments_from_db()
    executors = load_executors_from_db()
    
    if not executors or not tasks:
        return 0
    
    # Найти ID всех распределенных заявок
    assigned_task_ids = set(a['task_id'] for a in assignments)
    
    # Найти нераспределенные заявки
    unassigned_tasks = [t for t in tasks if t['id'] not in assigned_task_ids]
    
    if not unassigned_tasks:
        return 0
    
    assigned_count = 0
    
    # Распределить каждую нераспределенную заявку
    for task in unassigned_tasks:
        result = find_best_executor_simple(task, executors)
        if result:
            executor, score = result
            assignment = {
                'id': str(uuid.uuid4()),
                'task_id': task['id'],
                'executor_id': executor['id'],
                'assigned_at': datetime.now().isoformat(),
                'score': score
            }
            save_assignment_to_db(assignment)
            
            # Обновляем счетчик исполнителя
            executor['assigned_today'] += 1
            save_executor_to_db(executor)
            
            assigned_count += 1
            
            # Перезагружаем исполнителей для актуальных данных
            executors = load_executors_from_db()
    
    return assigned_count

def run_load_test_background(num_tasks, batch_size, delay_ms):
    """Фоновая функция для нагрузочного тестирования"""
    try:
        categories = ["IT", "Строительство", "Страхование", "Консалтинг"]
        priorities = ["Низкий", "Средний", "Высокий", "Критический"]
        
        set_load_test_status('running', 0, 0, num_tasks, 0)
        
        total_generated = 0
        total_assigned = 0
        start_time = time.time()
        
        for i in range(0, num_tasks, batch_size):
            # Проверяем статус (может быть остановлен пользователем)
            status = get_load_test_status()
            if status and status['status'] == 'stopped':
                break
            
            current_batch_size = min(batch_size, num_tasks - i)
            executors = load_executors_from_db()
            
            if not executors:
                set_load_test_status('error', message="Нет исполнителей!")
                return
            
            for j in range(current_batch_size):
                task_id = str(uuid.uuid4())
                category = random.choice(categories)
                priority = random.choice(priorities)
                
                # Генерируем параметры в зависимости от категории
                params = {}
                
                if category == "IT":
                    # IT параметры
                    all_skills = ["Python", "JavaScript", "React", "FastAPI", "Docker", "PostgreSQL", "AWS"]
                    params = {
                        'required_skills': random.sample(all_skills, random.randint(1, 3)),
                        'min_experience_years': random.choice([1, 2, 3, 5, 7]),
                        'complexity': random.randint(1, 10),
                        'remote_work': random.choice([True, False]),
                        'max_hourly_rate': random.choice([3000, 4000, 5000, 6000])
                    }
                
                elif category == "Строительство":
                    # Строительство параметры
                    params = {
                        'location': random.choice(["Москва", "Санкт-Петербург", "Казань", "Екатеринбург"]),
                        'equipment_needed': random.sample(["Кран", "Экскаватор", "Бетономешалка"], random.randint(1, 2)),
                        'square_meters': random.choice([500, 1000, 1500, 2000, 3000]),
                        'floor_count': random.randint(1, 10)
                    }
                
                elif category == "Страхование":
                    # Страхование параметры
                    params = {
                        'insurance_types': random.sample(["ОСАГО", "КАСКО", "Жизнь", "Имущество"], random.randint(1, 2)),
                        'vehicle_year': random.choice([2018, 2019, 2020, 2021, 2022, 2023]),
                        'driver_age': random.randint(25, 65),
                        'accident_history': random.choice([True, False])
                    }
                
                elif category == "Консалтинг":
                    # Консалтинг параметры
                    params = {
                        'required_certifications': random.sample(["PMP", "Agile", "PRINCE2", "Scrum Master"], random.randint(1, 2)),
                        'project_duration_months': random.choice([1, 3, 6, 12, 24]),
                        'team_size': random.randint(5, 50),
                        'industry': random.choice(["Финансы", "Производство", "Ритейл", "IT"])
                    }
                
                task = {
                    'id': task_id,
                    'name': f"Заявка #{total_generated + j + 1}",
                    'category': category,
                    'priority': priority,
                    'params': params,
                    'created_at': datetime.now().isoformat()
                }
                
                save_task_to_db(task)
                
                result = find_best_executor_simple(task, executors)
                if result:
                    executor, score = result
                    assignment = {
                        'id': str(uuid.uuid4()),
                        'task_id': task_id,
                        'executor_id': executor['id'],
                        'assigned_at': datetime.now().isoformat(),
                        'score': score
                    }
                    save_assignment_to_db(assignment)
                    
                    executor['assigned_today'] += 1
                    save_executor_to_db(executor)
                    
                    total_assigned += 1
            
            total_generated += current_batch_size
            
            # Обновляем прогресс в БД
            progress = total_generated / num_tasks
            set_load_test_status('running', progress, total_generated, num_tasks, total_assigned)
            
            if delay_ms > 0 and i + batch_size < num_tasks:
                time.sleep(delay_ms / 1000.0)
        
        end_time = time.time()
        elapsed_time = end_time - start_time
        performance = total_generated / elapsed_time if elapsed_time > 0 else 0
        
        # Завершение
        set_load_test_status('completed', 1.0, total_generated, num_tasks, total_assigned, elapsed_time, performance)
        
    except Exception as e:
        set_load_test_status('error', message=f"Ошибка: {str(e)}")

def render_load_test():
    st.markdown('<h2 class="section-header">🧪 Нагрузочное тестирование</h2>', unsafe_allow_html=True)
    
    # Получаем текущий статус теста из БД
    test_status_data = get_load_test_status()
    test_status = test_status_data['status'] if test_status_data else None
    
    if test_status == 'running':
        st.info("🔄 Тестирование выполняется в фоновом режиме. Вы можете переключаться между вкладками!")
        
        progress = test_status_data['progress']
        current = test_status_data['current']
        total = test_status_data['total']
        assigned = test_status_data['assigned']
        
        st.progress(progress)
        st.write(f"**Обработано:** {current}/{total} заявок | **Назначено:** {assigned}")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("⏹️ Остановить тестирование", type="secondary"):
                set_load_test_status('stopped')
                st.warning("Тестирование остановлено")
                st.rerun()
    
    with col2:
            if st.button("🔄 Обновить", type="secondary"):
                st.rerun()
    
    elif test_status == 'completed':
        elapsed = test_status_data['elapsed']
        performance = test_status_data['performance']
        current = test_status_data['current']
        assigned = test_status_data['assigned']
        
        st.success(f"""
✅ **Тестирование завершено!**  
Создано заявок: {current} | Назначено: {assigned}  
Время выполнения: {elapsed:.2f} сек | Производительность: {performance:.1f} заявок/сек
        """)
        
        if st.button("🔄 Запустить новое тестирование"):
            set_load_test_status('idle')
            st.balloons()
            st.rerun()
    
    elif test_status == 'error':
        error_msg = test_status_data['message']
        st.error(f"❌ Ошибка: {error_msg}")
        
        if st.button("🔄 Попробовать снова"):
            set_load_test_status('idle')
            st.rerun()
    
    # Форма настроек (показываем только если тест не запущен)
    if test_status not in ['running']:
        st.markdown("""
        ### 🎯 Массовая генерация заявок
        
        Этот инструмент позволяет сгенерировать большое количество заявок для демонстрации работы системы распределения под нагрузкой.
        **Тестирование работает в фоновом режиме - вы сможете переключаться между вкладками!**
        """)
    
    col1, col2 = st.columns(2)
    
    with col1:
        num_tasks = st.number_input("Количество заявок для генерации", min_value=10, max_value=10000, value=100, step=10)
        batch_size = st.number_input("Размер батча (заявок за раз)", min_value=1, max_value=100, value=10)
    
    with col2:
        delay_ms = st.slider("Задержка между батчами (мс)", min_value=0, max_value=1000, value=100, step=50)
    
    st.markdown("---")
    
    if st.button("🚀 Запустить нагрузочное тестирование", type="primary"):
        # Проверяем есть ли исполнители
        executors = load_executors_from_db()
        if not executors:
            st.error("❌ Нет исполнителей! Сначала добавьте исполнителей в разделе 'Исполнители'")
            return
        
        # Запускаем тестирование в отдельном потоке
        test_thread = threading.Thread(
            target=run_load_test_background,
            args=(num_tasks, batch_size, delay_ms),
            daemon=True
        )
        test_thread.start()
        
        st.success("✅ Нагрузочное тестирование запущено в фоновом режиме! Вы можете переключаться между вкладками.")
        time.sleep(0.5)  # Даем потоку запуститься
        st.rerun()
    
    # Статистика последнего теста
    st.markdown("### 📊 Текущая статистика")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Всего заявок в системе", len(st.session_state.tasks))
    
    with col2:
        st.metric("Всего назначений", len(st.session_state.assignments))
    
    with col3:
    active_executors = [e for e in st.session_state.executors if e.get('active', True)]
        total_assigned_today = sum(e['assigned_today'] for e in active_executors)
        st.metric("Всего назначено сегодня", total_assigned_today)

# Настройки
def render_settings():
    st.markdown('<h2 class="section-header">⚙️ Настройки системы</h2>', unsafe_allow_html=True)
    
    st.markdown("### 🎛 Общие настройки")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### 📊 Настройки дашборда")
        auto_refresh = st.checkbox("Автообновление дашборда", value=st.session_state.get('auto_refresh', True))
        st.session_state.auto_refresh = auto_refresh
    
        if auto_refresh:
            st.info("✅ Дашборд обновляется каждые 2 секунды")
        else:
            st.warning("⚠️ Автообновление отключено. Обновляйте страницу вручную")
    
    with col2:
        st.markdown("#### 🔄 Управление данными")
        
        if st.button("🔄 Сбросить дневные счетчики", type="secondary"):
            reset_daily_counts_in_db()
            st.session_state.executors = load_executors_from_db()
            st.success("✅ Дневные счетчики сброшены!")
            st.rerun()
    
    st.markdown("---")
    
    st.markdown("### 🗑️ Очистка данных")
    
    st.warning("⚠️ **Внимание!** Эти действия необратимы!")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("🗑️ Очистить все заявки и назначения", type="secondary"):
            clear_all_data_in_db()
            st.session_state.tasks = load_tasks_from_db()
            st.session_state.executors = load_executors_from_db()
            st.session_state.assignments = load_assignments_from_db()
            st.success("✅ Все заявки и назначения удалены!")
            st.rerun()
    
    with col2:
        if st.button("🗑️ Удалить всех исполнителей", type="secondary"):
            conn = get_sqlite_conn()
            cur = conn.cursor()
            cur.execute("DELETE FROM assignments")
            cur.execute("DELETE FROM executors")
            conn.commit()
            conn.close()
            st.session_state.executors = load_executors_from_db()
            st.session_state.assignments = load_assignments_from_db()
            st.success("✅ Все исполнители удалены!")
            st.rerun()
    
    st.markdown("---")
    
    st.markdown("### 📊 Информация о системе")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Версия БД", "SQLite 3")
    
    with col2:
        st.metric("Файл БД", os.path.basename(DB_PATH))
    
    with col3:
        db_size = os.path.getsize(DB_PATH) / 1024  # KB
        st.metric("Размер БД", f"{db_size:.1f} KB")

# Главная функция
def main():
    init_session_state()
    render_header()
    
    # Главное меню
    selected_page = render_main_menu()
    
    # Отображение выбранной страницы
    if selected_page == "dashboard":
        render_dashboard()
    elif selected_page == "executors":
        render_executors_management()
    elif selected_page == "load_test":
        render_load_test()
    elif selected_page == "settings":
        render_settings()

if __name__ == "__main__":
    main()
